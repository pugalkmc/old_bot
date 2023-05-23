import os
import re
import time
from datetime import datetime, timedelta
from typing import List

import openpyxl
import firebase_admin
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from firebase_admin import credentials, db
from openpyxl.styles import Alignment
from telegram import *
import telegram
from telegram.ext import *
from openpyxl.formula import Tokenizer
from openpyxl.utils.cell import get_column_letter
import os
from openpyxl import Workbook

cred = credentials.Certificate("kit-pro-f4b0d-firebase-adminsdk-mhzrf-8a07acab1c.json")
firebase_admin.initialize_app(cred, {
    "databaseURL": "https://kit-pro-f4b0d-default-rtdb.firebaseio.com/"
})

# Set up the Telegram bot

bot = Bot(token="6208523031:AAFfOb97T6Wml0pZUagE56A_MZDpCpUXZJk")


async def start(update, context):
    message = update.message
    chat_id = message.chat_id
    await bot.send_message(chat_id=chat_id,
                           text="Hi! I'm your Telegram bot. I'll collect messages and links from PoolSea Group")


async def collect_message(update, context):
    message = update.message
    username = message.from_user.username
    chat_id = message.chat_id
    group_id = message.chat.id
    chat_type = message.chat.type
    text = message.text

    member_list = ["srikanth084", "Jellys04", "Cryptomaker143",
                   "Shankar332", "Royce73", "Balaharishb",
                   "SaranKMC", "Sakthi_TVL"]

    if chat_type == "private":
        if "get " in message.text and len(text) > 6:
            text = text.replace("get ", "")
            await selva_sheet(update=update, context=context, date=text)
        elif username not in members:
            await bot.send_message(chat_id=chat_id, text="You have no permission to use this bot")

            return
        if "spreadsheet admin" in text:
            int_org = text.replace("spreadsheet", "")
            await save_to_spreadsheet(update, context, admin=int_org)

        elif "spreadsheet" in message.text and len(message.text) > 12:
            text = text.replace("spreadsheet ", "")
            await save_to_spreadsheet(update=update, context=context, date=text)


    elif chat_type == "group" or chat_type == "supergroup":
        # Only process messages from specific users in personal chat
        collection_name = datetime.now().strftime("%Y-%m-%d")
        message_id = message.message_id
        message_date_ist = (datetime.now() + timedelta(hours=5, minutes=30)).strftime(
            "%H:%M:%S")  # Convert datetime to IST timezone
        text = message.text
        if chat_id == -827109122:
            db.reference(f'selva/{collection_name}/{message_id}').set({
                'username': username,
                'text': text,
                'time': message_date_ist,
                'message_id': message_id
            })
        if username not in member_list or group_id != -1001588000922:
            return
        # Store message data in Firebase Realtime Database
        db.reference(f'messages/{collection_name}/{message_id}').set({
            'username': username,
            'text': text,
            'time': message_date_ist,
            'message_id': message_id
        })


admins_list = [1155684571, 814546021, 1291659507]


async def selva_sheet(update, context, admin=None, date=None):
    collection_name = date if date else datetime.now().strftime("%Y-%m-%d")
    await bot.send_message(chat_id=1292480260, text=f"request received for date: {collection_name}")

    messages = db.reference(f'selva/{collection_name}').get() or {}

    if len(messages) == 0 or messages is None:
        await bot.send_message(chat_id=1292480260, text="No message found for today")

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18

    ws['A1'] = 'Username'
    ws['B1'] = 'Message Text'
    ws['C1'] = 'IST Time'

    new_li = []
    msg_dict = {}
    row = 2
    for message_id, message_data in messages.items():
        username = message_data.get('username')
        text = message_data.get('text')
        time = message_data.get('time')
        new_li.append([username, text, time])
        ws.cell(row=row, column=1).value = username
        ws.cell(row=row, column=2).value = text
        ws.cell(row=row, column=3).value = time
        row += 1

        if username in msg_dict:
            msg_dict[username] += 1
        else:
            msg_dict[username] = 1
    wb.save(f"{collection_name}.xlsx")
    row = 1
    for i in msg_dict:
        ws.cell(row=row, column=5).value = i
        ws.cell(row=row, column=6).value = msg_dict[i]
        row += 1

    wb.save(f"{collection_name}.xlsx")

    await bot.send_document(chat_id=1292480260, document=open(f"{collection_name}.xlsx", "rb"))


async def save_to_spreadsheet(update=None, context=None, admin=None, date=None):
    collection_name = date if date else datetime.now().strftime("%Y-%m-%d")
    if update is None:
        chat_id = 1291659507
    else:
        chat_id = update.message.chat_id

    messages = db.reference(f'messages/{collection_name}').get() or {}
    if len(messages) == 0 or messages is None:
        await bot.send_message(chat_id=chat_id, text="Hey! ,No message found for today")

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20

    ws['A1'] = 'Username'
    ws['B1'] = 'Message Link'
    ws['C1'] = 'Message Text'
    ws['D1'] = 'IST Time'
    ws['F1'] = 'Username'
    ws['G1'] = 'Message Count'

    row = 2
    username_counts = {}
    for message_id, message_data in messages.items():
        username = message_data.get('username')
        text = message_data.get('text')
        time = message_data.get('time')
        link = f'https://t.me/poolsea/{message_id}'

        if username:
            if username in username_counts:
                username_counts[username]['count'] += 1
            else:
                username_counts[username] = {'count': 1, 'total': 0}

            ws.cell(row=row, column=1).value = username
            ws.cell(row=row, column=2).value = link
            ws.cell(row=row, column=3).value = text
            ws.cell(row=row, column=4).value = time
            row += 1
    msg = ""
    for i in username_counts:
        msg += f"{i} - {username_counts[i]['count']}\n"
    await bot.send_message(chat_id=update.message.chat_id, text=f"Total Messages: {len(messages.items())}\n\n{msg}")

    if admin is not None and "group" in admin:
        await bot.send_message(chat_id=-1001586628789,
                               text=f"Count for {collection_name}: {len(messages.items())}\n\n{msg}")

    ws["F1"] = "Usernames"
    ws["G1"] = "Count"

    member_list = ["srikanth084", "Jellys04", "Cryptomaker143",
                   "Shankar332", "Royce73", "Balaharishb",
                   "SaranKMC", "Sakthi_TVL"]

    index = 0
    for row in range(2, len(member_list) + 2):
        count = '=COUNTIF(A:A,"*' + member_list[index] + '*")'
        ws.cell(row=row, column=6).value = member_list[index]
        ws.cell(row=row, column=7).value = count
        index += 1

    file_name = f"{collection_name}.xlsx"
    wb.save(file_name)
    if admin is None:
        await bot.send_document(chat_id=chat_id, document=open(file_name, 'rb'))
    if admin is not None and "admin" in admin:
        if "group" in admin:
            await bot.send_document(chat_id=-1001586628789, document=open(file_name, "rb"))
        for i in admins_list:
            await bot.send_document(chat_id=i, document=open(file_name, "rb"))


def run_spreadsheet_job():
    save_to_spreadsheet(admin="group")


def schedule_spreadsheet_job():
    scheduler = AsyncIOScheduler()
    scheduler.add_job(run_spreadsheet_job, 'cron', hour=14, minute=52, second=0, timezone='Asia/Kolkata')
    scheduler.start()


BOT_TOKEN = "6208523031:AAFfOb97T6Wml0pZUagE56A_MZDpCpUXZJk"


def main():
    dp = Application.builder().token(BOT_TOKEN).build()
    import threading
    threading.Thread(target=schedule_spreadsheet_job).start()
    dp.add_handler(CommandHandler("start", start))
    dp.add_handler(CommandHandler("spreadsheet", save_to_spreadsheet))
    dp.add_handler(MessageHandler(filters.TEXT, collect_message))
    dp.run_polling()


if __name__ == '__main__':
    main()


